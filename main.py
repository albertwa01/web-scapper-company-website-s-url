
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import time


def search_company(company_name,driver):
    
    search_box = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.NAME, "q")))
    search_box.send_keys(company_name+" Website")
    search_box.submit()

    SCROLL_PAUSE_TIME = 5
    last_height = driver.execute_script("return document.body.scrollHeight")

    while True:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(SCROLL_PAUSE_TIME)
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height
    links = []
    search_result_links = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "h2 a")))#driver.find_elements_by_css_selector("h2 a")
    urls = [link.get_attribute("href") for link in search_result_links]
    # print(urls)
    urlss= []
    for url in urls:
        # print(url)
        if "https://www.bing.com" not in url:
            urlss.append(url)
    driver.get("https://www.bing.com/")

    return urlss[:3]

from openpyxl import *
def main_func():
    chrome_options = Options()
    chrome_options.add_argument('Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36')
    driver = webdriver.Chrome(options=chrome_options)
    driver.get("https://www.bing.com/")
    excel_path=r"G:\FR\web_scraping\Website-Sample.xlsx"
    wb_in=load_workbook(excel_path,data_only=True)
    ws_in=wb_in.active

    wb=Workbook()
    ws=wb.active
    ws.append(["Company Name","Match 1","Match 2","Match 3"])
    print(ws_in.max_row+1,"ws_in.max_row+1")
    for i in range(2,ws_in.max_row+1):
    # for i in range(51,1,-1):
        company_name=ws_in['A'+str(i)].value
        print(company_name)
        matches = search_company(company_name,driver)
        print(company_name,"__"*10,matches)
        out_li=[company_name]
        for mat in matches:
            out_li.append(mat)
        ws.append(out_li)
        wb.save(r"G:\FR\web_scraping\Website-Sample-output_top3.xlsx")
    wb.save(r"G:\FR\web_scraping\Website-Sample-output_top3.xlsx")


    # Quit the browser
    driver.quit()


main_func()
# search_company("sequelstring")
